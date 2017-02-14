#   Copyright 2017 CSIRO
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.


import datetime
import csv
from decimal import Decimal, ROUND_HALF_UP
import re
from openpyxl import load_workbook
from django.db.models import Sum
from dashboard_loader.loader_utils import *
from coag_uploader.models import *
from indigenous_naplan_uploader.models import *
from coag_uploader.uploader import load_state_grid, load_benchmark_description, update_graph_data, populate_raw_data, populate_crosstab_raw_data, update_stats, update_state_stats, column_labels
from widget_def.models import Parametisation



# These are the names of the groups that have permission to upload data for this uploader.
# If the groups do not exist they are created on registration.
groups = [ "upload_all", "upload" ]

# This describes the file format.  It is used to describe the format by both
# "python manage.py upload_data frontlineservice_uploader" and by the uploader 
# page in the data admin GUI.


file_format = {
    "format": "xlsx",
    "sheets": [
            {
                "name": "Data",
                "cols": [ 
                            ('A', 'Value ("Proportion at or above NMS (%)", "95% confidence interval for proportion (% points)", "Trajectory point for the proportion (%)")'),
                            ('B', 'Domain/Subject ("Reading", "Numeracy")'),
                            ('C', 'Year level (3, 5 ,7 ,9)'),
                            ('...', 'Column per state plus Australia'),
                        ],
                "rows": [
                            ('1', "Heading row"),
                            ('...', 'Set of rows per combination of subject/domain and year level, with one row per value (and possibly other lines which are ignored)',)
                        ],
                "notes": [
                    'Blank rows and columns ignored',
                ],
            },
            {
                "name": "Description",
                "cols": [
                            ('A', 'Key'),
                            ('B', 'Value'),
                        ],
                "rows": [
                            ('Measure', 'Full description of benchmark'),
                            ('Short Title', 'Short widget title (not used)'),
                            ('Status', 'Benchmark status'),
                            ('Updated', 'Year data last updated'),
                            ('Desc body', 'Body of benchmark status description. One paragraph per line.'),
                            ('Notes', 'Notes for benchmark status description.  One note per line.'),
                        ],
                "notes": [
                         ],
            }
        ],
}

def upload_file(uploader, fh, actual_freq_display=None, verbosity=0):
        messages = []
#    try:
        if verbosity > 0:
            messages.append("Loading workbook...")
        wb = load_workbook(fh, read_only=True)
        desc = load_benchmark_description(wb, "Description")
        messages.extend(
                load_naplan_grid(wb, "Data",
                                "Indigenous", "Naplan",
                                IndigenousNaplanData,
                                {
                                    "indig_proportion_above_nms": "Proportion at or above NMS (%)",
                                    "indig_confidence_interval": "95% confidence interval for proportion (% points)",
                                    "indig_trajectory": "Trajectory point for the proportion(%)",
                                },
                                "Values", "Domain", "Year level",
                                verbosity=verbosity)
        )
        messages.extend(update_stats(desc, None,
                            "indig_lit-indigenous-hero", "indig_lit-indigenous-hero", 
                            "indig_lit-indigenous-hero-state", "indig_lit-indigenous-hero-state", 
                            "indigenous_indig_lit", "indigenous_indig_lit", 
                            "indigenous_indig_lit_state", "indigenous_indig_lit_state", 
                            verbosity))
        messages.extend(update_stats(desc, None,
                            "indig_num-indigenous-hero", "indig_num-indigenous-hero", 
                            "indig_num-indigenous-hero-state", "indig_num-indigenous-hero-state", 
                            "indigenous_indig_num", "indigenous_indig_num", 
                            "indigenous_indig_num_state", "indigenous_indig_num_state", 
                            verbosity))
        messages.extend(update_state_stats(
                            "indig_lit-indigenous-hero-state", "indig_lit-indigenous-hero-state", 
                            "indigenous_indig_lit_state", "indigenous_indig_lit_state", 
                            IndigenousNaplanData, [],
                            combine_all=True,
                            use_benchmark_tls=True,
                            query_filter_kwargs={
                                "subject": READING
                            },
                            status_func=IndigenousNaplanData.tlc,
                            verbosity=verbosity))
        messages.extend(update_state_stats(
                            "indig_num-indigenous-hero-state", "indig_num-indigenous-hero-state", 
                            "indigenous_indig_num_state", "indigenous_indig_num_state", 
                            IndigenousNaplanData, [],
                            combine_all=True,
                            use_benchmark_tls=True,
                            query_filter_kwargs={
                                "subject": NUMERACY
                            },
                            status_func=IndigenousNaplanData.tlc,
                            verbosity=verbosity))
        for obj in IndigenousNaplanData.objects.filter(subject=READING, state=AUS):
            set_statistic_data("indig_lit-indigenous-hero", "indig_lit-indigenous-hero", 
                            "year_%d" % obj.year_lvl, "Year %d" % obj.year_lvl,
                            traffic_light_code=obj.tlc())
            set_statistic_data("indigenous_indig_lit", "indigenous_indig_lit", 
                            "year_%d" % obj.year_lvl, "Year %d" % obj.year_lvl,
                            traffic_light_code=obj.tlc())
        for obj in IndigenousNaplanData.objects.filter(subject=READING):
            set_statistic_data("indigenous_indig_lit", "indigenous_indig_lit", 
                            "%s_year_%d" % (obj.state_display().lower(), obj.year_lvl), 
                            obj.indig_proportion_above_nms,
                            traffic_light_code=obj.tlc())
        for obj in IndigenousNaplanData.objects.filter(subject=NUMERACY, state=AUS):
            set_statistic_data("indig_num-indigenous-hero", "indig_num-indigenous-hero", 
                            "year_%d" % obj.year_lvl, "Year %d" % obj.year_lvl,
                            traffic_light_code=obj.tlc())
            set_statistic_data("indigenous_indig_num", "indigenous_indig_num", 
                            "year_%d" % obj.year_lvl, "Year %d" % obj.year_lvl,
                            traffic_light_code=obj.tlc())
        for obj in IndigenousNaplanData.objects.filter(subject=NUMERACY):
            set_statistic_data("indigenous_indig_num", "indigenous_indig_num", 
                            "%s_year_%d" % (obj.state_display().lower(), obj.year_lvl), 
                            obj.indig_proportion_above_nms,
                            traffic_light_code=obj.tlc())
        messages.extend(
                populate_raw_data(
                            "indigenous_indig_num", "indigenous_indig_num", 
                            "indigenous_indig_num", IndigenousNaplanData, 
                                {
                                    "indig_proportion_above_nms": "proportion_above_nms",
                                    "indig_confidence_interval": "confidence_interval",
                                    "indig_trajectory": "trajectory_point"
                                },
                            query_kwargs={ "subject": NUMERACY },
                            use_dates=False)
                )
        messages.extend(
                populate_raw_data(
                            "indigenous_indig_lit", "indigenous_indig_lit", 
                            "indigenous_indig_lit", IndigenousNaplanData, 
                                {
                                    "indig_proportion_above_nms": "proportion_above_nms",
                                    "indig_confidence_interval": "confidence_interval",
                                    "indig_trajectory": "trajectory_point"
                                },
                            query_kwargs={ "subject": READING },
                            use_dates=False)
                )
        p = Parametisation.objects.get(url="state_param")
        for pval in p.parametisationvalue_set.all():
            state_num = state_map[pval.parameters()["state_abbrev"]]
            for obj in IndigenousNaplanData.objects.filter(subject=READING, state=state_num):
                set_statistic_data("indig_lit-indigenous-hero-state", "indig_lit-indigenous-hero-state", 
                                "year_%d" % obj.year_lvl, "Year %d" % obj.year_lvl,
                                traffic_light_code=obj.tlc(),
                                pval=pval)
                set_statistic_data("indigenous_indig_lit_state", "indigenous_indig_lit_state", 
                                "year_%d" % obj.year_lvl, "Year %d" % obj.year_lvl,
                                traffic_light_code=obj.tlc(),
                                pval=pval)
                set_statistic_data("indigenous_indig_lit_state", "indigenous_indig_lit_state", 
                                "state_year_%d" % obj.year_lvl, 
                                obj.indig_proportion_above_nms,
                                traffic_light_code=obj.tlc(),
                                pval=pval)
            for obj in IndigenousNaplanData.objects.filter(subject=READING, state=AUS):
                set_statistic_data("indigenous_indig_lit_state", "indigenous_indig_lit_state", 
                                "australia_year_%d" % obj.year_lvl, 
                                obj.indig_proportion_above_nms,
                                traffic_light_code=obj.tlc(),
                                pval=pval)
            for obj in IndigenousNaplanData.objects.filter(subject=NUMERACY, state=state_num):
                set_statistic_data("indig_num-indigenous-hero-state", "indig_num-indigenous-hero-state", 
                                "year_%d" % obj.year_lvl, "Year %d" % obj.year_lvl,
                                traffic_light_code=obj.tlc(),
                                pval=pval)
                set_statistic_data("indigenous_indig_num_state", "indigenous_indig_num_state", 
                                "year_%d" % obj.year_lvl, "Year %d" % obj.year_lvl,
                                traffic_light_code=obj.tlc(),
                                pval=pval)
                set_statistic_data("indigenous_indig_num_state", "indigenous_indig_num_state", 
                                "state_year_%d" % obj.year_lvl, 
                                obj.indig_proportion_above_nms,
                                traffic_light_code=obj.tlc(),
                                pval=pval)
            for obj in IndigenousNaplanData.objects.filter(subject=NUMERACY, state=AUS):
                set_statistic_data("indigenous_indig_num_state", "indigenous_indig_num_state", 
                                "australia_year_%d" % obj.year_lvl, 
                                obj.indig_proportion_above_nms,
                                traffic_light_code=obj.tlc(),
                                pval=pval)
            messages.extend(
                    populate_raw_data(
                                "indigenous_indig_num_state", "indigenous_indig_num_state", 
                                "indigenous_indig_num", IndigenousNaplanData, 
                                    {
                                        "indig_proportion_above_nms": "proportion_above_nms",
                                        "indig_confidence_interval": "confidence_interval",
                                        "indig_trajectory": "trajectory_point"
                                    },
                                query_kwargs={ "subject": NUMERACY },
                                use_dates=False,
                                pval=pval)
                    )
            messages.extend(
                    populate_raw_data(
                                "indigenous_indig_lit_state", "indigenous_indig_lit_state", 
                                "indigenous_indig_lit", IndigenousNaplanData, 
                                    {
                                        "indig_proportion_above_nms": "proportion_above_nms",
                                        "indig_confidence_interval": "confidence_interval",
                                        "indig_trajectory": "trajectory_point"
                                    },
                                query_kwargs={ "subject": READING },
                                use_dates=False,
                                pval=pval)
                    )
#    except LoaderException, e:
#        raise e
#    except Exception, e:
#        raise LoaderException("Invalid file: %s" % unicode(e))
        return messages

def read_cell(sheet, col, row, strip=False):
    cval = sheet["%s%d" % (col, row)].value
    if cval is not None and strip:
        cval = cval.strip()
    return cval

def reset_objs(states, subject, year_level):
    return { st: {
                    "state": state_map[st], 
                    "year_lvl": year_level, 
                    "subject": subject_map[subject],
                } for st in states }

def load_naplan_grid(wb, sheet_name,
                                data_category, dataset, model,
                                cell_rows, row_discrim_heading, subject_heading, year_level_heading,
                                verbosity=0):
    messages = []
    discriminators = { v:k for k ,v in cell_rows.items() }
    if verbosity > 2:
        messages.append("Loading %s Data: %s" % (data_category, dataset))
    sheet = wb[sheet_name]
    row = 1
    records_written = 0
    subject_col=None
    year_level_col=None
    rowdiscrim_col=None
    states=[
        'nsw', 'vic', 'qld', 'wa', 'sa', 'tas', 'act', 'nt', 'aust',
    ]
    state_cols = {}
    columns_mapped = False
    model.objects.all().delete()
    kwargs = {}
    subj=None
    year_level=None
    objs = {}
    while True:
        if not columns_mapped:
            for col in column_labels(1, sheet.max_column):
                try:
                    cval = read_cell(sheet, col, row, True)
                except IndexError:
                    raise LoaderException("Sheet finished, column headings not found")
                if cval.lower() == subject_heading.lower():
                    subject_col = col
                if cval.lower() == year_level_heading.lower():
                    year_level_col = col
                if cval.lower() == row_discrim_heading.lower():
                    rowdiscrim_col = col
                for st in states:
                    if cval.lower() == st.lower():
                        state_cols[st] = col
            columns_mapped = True
            for st in states:
                if st not in state_cols:
                    columns_mapped = False
                    break
            if not subject_col or not year_level_col or not rowdiscrim_col:
                columns_mapped = False
        else:
            try:
                _subj = read_cell(sheet, subject_col, row, True)
                _yrlvl= read_cell(sheet, year_level_col, row)
                if _subj != subj or _yrlvl != year_level:
                    if subj and year_level:
                        # Write out accumulated objects
                        for obj in objs.values():
                            o = model(**obj)
                            o.save()
                            records_written += 1
                    # Reset accumulated objects
                    subj = _subj
                    year_level = _yrlvl
                    objs = reset_objs(states, subj, year_level) 
                # Accumulate objects
                cval = read_cell(sheet, rowdiscrim_col, row, True)
                if cval in discriminators:
                    fldname = discriminators[cval]
                    for st, scol in state_cols.items():
                        objs[st][fldname]=read_cell(sheet, scol, row)
            except IndexError:
                if subj and year_level:
                    # Write out accumulated objects
                    for obj in objs.values():
                        o = model(**obj)
                        o.save()
                        records_written += 1
                if verbosity > 1:
                    messages.append("Records written: %d" % records_written)
                return messages
        row += 1
