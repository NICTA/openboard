#   Copyright 2016,2017 NICTA
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.


import datetime
import csv
import math
from scipy import stats
from decimal import Decimal, ROUND_HALF_UP
import re
from openpyxl import load_workbook
from dashboard_loader.loader_utils import *
from widget_def.models import Parametisation, ParametisationValue
from coag_uploader.models import *
from django.template import Template, Context

def column_labels(mini, maxi):
    labels=[]
    alphabet = [ chr(i) for i in range(ord('A'), ord('A') + 26) ]
    alphabetp = [''] + alphabet
    ri = 0
    for c0 in alphabetp:
        for c1 in alphabet:
            ri += 1
            col = c0 + c1
            if ri >= mini and ri <= maxi:
                labels.append(col)
            if ri > maxi:
                return labels

def find_state_cols(sheet, row):
    state_cols = {}
    try:
        for col in column_labels(sheet.min_column, sheet.max_column):
           val = sheet["%s%d" % (col, row)].value
           if val in state_map:
                state_cols[state_map[val]] = col
    except IndexError:
           pass
    return state_cols

def zero_all_rows(rows):
    for k in rows.keys():
        rows[k] = 0

def all_rows_found(rows, optional_rows, defaults):
    got_any = False
    for k in rows.keys():
        if not rows[k] and k not in optional_rows and k not in defaults:
            return False
        if rows[k]:
            got_any=True
    return got_any

def listify(x):
    if isinstance(x, basestring):
        return [ x, ]
    else:
        return x

def load_state_grid(wb, sheet_name, data_category, dataset, abort_on, model, first_cell_rows, intermediate_cell_rows, verbosity=0, transforms={}, fld_defaults={}, use_dates=True, multi_year=False, date_field=None, date_parser=None, optional_rows=[]):
    messages = []
    if first_cell_rows:
        raise LoaderException("first_cell_rows no longer supported")
    if verbosity > 2:
        messages.append("Loading %s Data: %s" % (data_category, dataset))
    sheet = wb[sheet_name]
    state_cols = {}
    row = 1
    while not state_cols:
        state_cols = find_state_cols(sheet, row)
        row += 1
        if not state_cols and row > sheet.max_row:
            raise LoaderException("No State Columns found in worksheet '%s'" % sheet_name)
    date = None
    isfy = False
    myr = 1
    _isfy = False
    _myr = 1
    rows = {}
    records_written = 0
    for fld in intermediate_cell_rows.keys():
        rows[fld] = 0
    if use_dates:
        start_from_col = 2
    else:
        start_from_col = 1
    def write_record(rows, date, isfy, myr):
        recs_written = 0
        for state, scol in state_cols.items():
            nulldata = False
            if use_dates and not date_field:
                defaults = { "financial_year": isfy, "multi_year": myr }
            else:
                defaults = {}
            for def_fld, def_val in fld_defaults.items():
                defaults[def_fld] = def_val
            for fld, frow in rows.items():
                if not frow:
                    continue
                rawval = sheet["%s%d" % (scol, frow)].value
                if fld in transforms:
                    rawval = transforms[fld](rawval)
                if rawval is not None or fld not in defaults:
                    defaults[fld] = rawval
                if defaults[fld] is None and fld not in fld_defaults and fld not in transforms:
                    nulldata = True
            kwargs = { "state": state }
            if date_field:
                kwargs[date_field] = date
            elif use_dates:
                kwargs["year"] = date
            if nulldata:
                delresult=model.objects.filter(**kwargs).delete()
                if delresult[0] > 0 and verbosity > 0:
                    messages.append("Deleted %d objects: %s" % (delresult[0], repr(delresult[1])))
                continue
            kwargs["defaults"] = defaults
            obj, created = model.objects.update_or_create(**kwargs)
            recs_written += 1
        zero_all_rows(rows)
        return recs_written
    blank_rows = 0
    while True:
        try:
            first_cell=sheet["A%d" % row].value
        except IndexError:
            if all_rows_found(rows, optional_rows, fld_defaults):
                records_written += write_record(rows, date, isfy, myr)
            else:
                zero_all_rows(rows)
            break
        if first_cell:
            if isinstance(first_cell, unicode):
                first_cell = first_cell.strip()
            if first_cell == abort_on:
                break
            if use_dates:
                if date_parser:
                    _date = date_parser(first_cell)
                    if _date and _date != date:
                        if all_rows_found(rows, optional_rows, fld_defaults):
                            records_written += write_record(rows, date, None, None)
                        else:
                            zero_all_rows(rows)
                        date = _date
                try:
                    if multi_year:
                        (_year, _myr) = parse_multiyear(first_cell)
                    else:
                        (_year, _isfy) = parse_year(first_cell)
                    new_year = False
                    if _year != date:
                        new_year = True
                    if _myr != myr:
                        new_year = True
                    if _isfy != isfy:
                        new_year = True
                    if new_year:
                        if all_rows_found(rows, optional_rows, fld_defaults):
                            records_written += write_record(rows, date, isfy, myr)
                        else:
                            zero_all_rows(rows)
                        date = _year
                        isfy = _isfy
                        myr = _myr
                except Exception, e:
                    pass
        else:
            blank_rows += 1
            if blank_rows > 100:
                if all_rows_found(rows, optional_rows, fld_defaults):
                    records_written += write_record(rows, date, isfy, myr)
                else:
                    zero_all_rows(rows)
                break
        matches = []
        if date or not use_dates:
            for col in column_labels(start_from_col, sheet.max_column):
                if col in state_cols.values():
                    break
                cval = sheet["%s%d" % (col, row)].value
                if isinstance(cval, unicode):
                    cval = cval.strip()
                for fld, icr in intermediate_cell_rows.items():
                    icr = listify(icr)
                    for icri in icr:
                        if cval == icri:
                            matches.append(icri)
                            break
                    full_match = True
                    for icri in icr:
                        if icri not in matches:
                            full_match = False
                            break
                    if full_match:
                        rows[fld]=row
                        break
        row += 1
    if verbosity > 1:
        messages.append("Records written: %d" % records_written)
    return messages

def load_progress_grid(wb, sheet_name, data_category, dataset, model,
                                cell_rows, transforms={}, verbosity=0):
    messages = []
    if verbosity > 2:
        messages.append("Loading %s Data: %s" % (data_category, dataset))
    sheet = wb[sheet_name]
    row = 1
    records_written = 0
    column_map = {}
    columns_mapped = False
    model.objects.all().delete()
    while True:
        if not columns_mapped:
            for col in column_labels(1, sheet.max_column):
                try:
                    cval = sheet["%s%d" % (col, row)].value
                    if cval is not None:
                        cval = cval.strip()
                except IndexError:
                    raise LoaderException("Sheet finished, column headings not found")
                for k, v in cell_rows.items():
                    if cval == v:
                        column_map[k] = col
            columns_mapped = True
            for col in cell_rows.keys():
                if col not in column_map:
                    columns_mapped = False
                    break
        else:
            kwargs = {}
            for fldname, col in column_map.items():
                try:
                    cval = sheet["%s%d" % (col, row)].value
                    if not cval:
                        raise IndexError()
                except IndexError:
                    if verbosity > 1:
                        messages.append("Records written: %d" % records_written)
                    return messages
                if cval is None:
                    break
                cval = cval.strip()
                if fldname in transforms:
                    cval = transforms[fldname](cval)
                kwargs[fldname] = cval
            if kwargs:
                obj = model(**kwargs)
                obj.save()
                records_written += 1
        row += 1

def flt_year_equal(y1,y2):
    return abs(y1-y2) < 1e-6

def current_float_year():
    today=datetime.date.today()
    flt_year = float(today.year)
    if today.month >= 7:
        flt_year += 0.5
    return flt_year

def find_trend(metrics, reference_year, target_year):
    reference = None
    datum = None
    years = []
    data = []
    for metric in metrics:
        years.append(metric[0])
        data.append(metric[1])
        if flt_year_equal(metric[0], reference_year):
            reference = metric[1]
        if flt_year_equal(metric[0], target_year):
            datum = metric[1]
    if reference is None:
        return None
    if datum is None:
        slope, intercept, r_val, p_val, std_err = stats.linregress(years,data)
        datum = slope * target_year + intercept
        projected = True
    else:
        projected = False
    return {
            "datum": datum, 
            "reference": reference, 
            "benchmark": (datum-reference)/reference,
            "projected": projected
    }

TARGET_RELATIVE_CHANGE = 1
TARGET_ABSOLUTE_CHANGE = 2
TARGET_ABSOLUTE_VALUE  = 3

def benchmark_met(target, trend,
                desire_overtarget, target_type):
    if target_type == TARGET_RELATIVE_CHANGE:
        return (desire_overtarget and trend["benchmark"] >= target) or (not desire_overtarget and trend["benchmark"] <= target)
    elif target_type == TARGET_ABSOLUTE_CHANGE:
        return (desire_overtarget and trend["datum"]-trend["reference"] >= target) or (not desire_overtarget and trend["datum"]-trend["reference"] <= target)
    else:
        return (desire_overtarget and trend["datum"] >= target) or (not desire_overtarget and trend["datum"] <= target)

def calculate_benchmark(reference_year, target_year, 
                    target, desire_overtarget, 
                    model, benchmark_field, 
                    widget_url, widget_label, 
                    verbosity, target_type=TARGET_RELATIVE_CHANGE):
    messages = []
    metrics = [ (obj.float_year(), float(getattr(obj,benchmark_field))) for obj in model.objects.filter(state=AUS).order_by("year", "financial_year") ]
    trend = find_trend(metrics, reference_year, target_year)
    if trend:
        if benchmark_met(target, trend, desire_overtarget, target_type):
            if trend["projected"]:
                if target_year >= current_float_year():
                    tlc = "likely_to_have_been_met"
                else:
                    tlc = "on_track"
            else:
                tlc = "achieved"
        else:
            if trend["projected"]:
                if target_year >= current_float_year():
                    tlc = "unlikely_to_have_been_met"
                else:
                    tlc = "not_on_track"
            else:
                tlc = "not_met"
        if trend["projected"]:
            outcome_type = "projection"     
        else:
            outcome_type = "complete"
        if target_type == TARGET_RELATIVE_CHANGE:
            benchmark = trend["benchmark"]*100.0
        elif target_type == TARGET_ABSOLUTE_CHANGE:
            benchmark = trend["datum"] - datum["reference"]
        else:
            benchmark = trend["datum"]
        datum = trend["datum"]
        reference = trend["reference"]
    else:
        benchmark = 0.0
        reference = 0.0
        datum = 0.0
        outcome_type = "N/A"
        tlc = "new_revised_benchmark"
    set_statistic_data(widget_url, widget_label, "outcome", 
                    abs(benchmark), traffic_light_code=tlc, 
                    trend=cmp(datum, reference))
    set_statistic_data(widget_url, widget_label, "outcome_type", 
                    outcome_type)
    set_statistic_data(widget_url, widget_label, "benchmark", 
                    abs(target*100.0), trend=cmp(target, 0.0))
    set_statistic_data(widget_url, widget_label, "benchmark_year", 
                    display_float_year(target_year))
    clear_statistic_list(widget_url, widget_label, "data")
    add_statistic_list_item(widget_url, widget_label, "data",
                            reference, 10,
                            label=display_float_year(reference_year))
    last_metric = metrics[-1]
    add_statistic_list_item(widget_url, widget_label, "data",
                            last_metric[1], 20,
                            label=display_float_year(last_metric[0]))
    return messages

def calculate_indicator(desire_increase, 
                    model, indicator_field, 
                    widget_url, widget_label, 
                    verbosity):
    messages = []
    metrics = [ (obj.float_year(), float(getattr(obj,indicator_field))) for obj in model.objects.filter(state=AUS).order_by("year", "financial_year") ]
    if desire_increase:
        indicator_trend = 1
    else:
        indicator_trend = -1
    if len(metrics) == 0:
        # New indicator
        trend = 0
        tlc = "mixed_results_or_new_indicator"
        outcome_years = "N/A"
        old = None
        new = None
    else:
        if len(metrics) == 1:
            old = metrics[0]
            new = None
            trend = 0
            outcome_years = "%s" % display_float_year(old[0])
        else:
            old = metrics[0]
            new = metrics[-1]
            trend = cmp(new[1],old[1])
            outcome_years = "%s - %s" % (display_float_year(old[0]), display_float_year(new[0]))
        if trend == 0:
            tlc = "no_improvement"
        elif (trend > 0 and desire_increase) or (trend < 0 and not desire_increase):
            tlc = "improving"
        else:
            tlc = "negative_change"
    set_statistic_data(widget_url, widget_label, "outcome", 
                    "", traffic_light_code=tlc, 
                    trend=trend)
    set_statistic_data(widget_url, widget_label, "outcome_years", 
                    outcome_years)
    set_statistic_data(widget_url, widget_label, "indicator", 
                    "", trend=indicator_trend)
    clear_statistic_list(widget_url, widget_label, "data")
    if old:
        add_statistic_list_item(widget_url, widget_label, "data",
                    old[1], 10, label=display_float_year(old[0]))
    if new:
        add_statistic_list_item(widget_url, widget_label, "data",
                    new[1], 20, label=display_float_year(new[0]))
    return messages

benchmark_statuses = {
    "achieved": {
        "short": "Achieved",
        "long": "The final assessment date for the benchmark has passed. The benchmark was met.",
        "tlc": "achieved",
        "icon": "yes",
    },
    "on_track": {
        "short": "On track",
        "long": "The final assessment date for this benchmark is in the future. On the basis of results so far, the benchmark is on track to be met.",
        "tlc": "on_track",
        "icon": "yes",
    },
    "likely_to_have_been_met": {
        "short": "Likely to have been met",
        "long": "The due date for the benchmark assessment has passed, but we do not yet have data to assess results at that date. On the basis of the available data, the benchmark is likely to have been met.",
        "tlc": "likely_to_have_been_met",
        "icon": "yes",
    },
    "unlikely_to_have_been_met": {
        "short": "Unlikely to have been met",
        "long": "The due date for the benchmark assessment has passed, but we do not yet have data to assess results at that date. On the basis of the available data, the benchmark is unlikely to have been met.",
        "tlc": "unlikely_to_have_been_met",
        "icon": "warning",
    },
    "not_on_track": {
        "short": "Not on track",
        "long": "The final assessment date for this benchmark is in the future. On the basis of results so far, the benchmark is not on track to be met.",
        "tlc": "on_track",
        "icon": "warning",
    },
    "not_met": {
        "short": "Not met",
        "long": "The final assessment date for the benchmark is has passed. The benchmark was not met.",
        "tlc": "not_met",
        "icon": "no",
    },
    "new_benchmark": {
        "short": "Baseline for target",
        "long": "There is no time series data available for this benchmark yet, so it is not possible to assess progress at this point.",
        "tlc": "new_benchmark",
        "icon": "unknown",
    },
    "revised_benchmark": {
        "short": "Revised benchmark",
        "long": "An agreement has been reached to replace a previous benchmark.",
        "tlc": "revised_benchmark",
        "icon": "unknown",
    },
    "mixed_results": {
        "short": "Mixed Results",
        "long": "This benchmark includes a suite of results, which have shown a variety of positive, negative and/or no change. It is not possible to form an overall traffic light assessment.",
        "tlc": "mixed_results",
        "icon": "unknown",
    },
    "no_trajectory": {
        "short": "No Trajectory",
        "long": "This jurisdiction has not agreed to a trajectory to support this target.",
        "tlc": "not_applicable",
        "icon": "unknown",
    },
}
indicator_statuses = {
    "improving": {
        "short": "Improving",
        "long": "There has been a noticable improvement on this measure.",
        "tlc": "improving",
        "incr_trend": 1,
        "decr_trend": -1,
        "icon": "yes",
    },
    "no_improvement": {
        "short": "No Improvement",
        "long": "There has been no noticable change across this measure.",
        "tlc": "no_improvement",
        "incr_trend": 0,
        "decr_trend": 0,
        "icon": "warning",
    },
    "negative_change": {
        "short": "Negative change",
        "long": "There has been a noticable worsening on this measure.",
        "tlc": "negative_change",
        "incr_trend": -1,
        "decr_trend": 1,
        "icon": "no",
    },
    "mixed_results": {
        "short": "Mixed results",
        "long": "This indicator includes a suite of results, which have shown a variety of positive, negative and/or no change. It is not possible to form an overall traffic light assessment.",
        "tlc": "mixed_results",
        "incr_trend": 0,
        "decr_trend": 0,
        "icon": "unknown",
    },
    "new_indicator": {
        "short": "New indicator",
        "long": "There is no time series data available for this indicator yet, so it is not possible to assess progress at this point.",
        "tlc": "new_indicator",
        "incr_trend": 0,
        "decr_trend": 0,
        "icon": "unknown",
    },
    "no_data": {
        "short": "No data",
        "long": "There is no data available for this indicator, so it is not possible to assess progress.",
        "tlc": "no_data",
        "incr_trend": 0,
        "decr_trend": 0,
        "icon": "unknown",
    },
    "no_trend_data": {
        "short": "No trend data",
        "long": "This jurisdiction has not agreed to a trajectory to support this target.",
        "tlc": "no_trend_data",
        "incr_trend": 0,
        "decr_trend": 0,
        "icon": "unknown",
    },
    "no_trajectory": {
        "short": "No trajectory",
        "long": "There is no time-series data available for this indicator, so it is not possible to assess progress.",
        "tlc": "no_trend_data",
        "incr_trend": 0,
        "decr_trend": 0,
        "icon": "unknown",
    },
}

def load_benchmark_description(wb, sheetname, indicator=False, additional_lookups={}):
    key_lookup = {
        "measure": "measure",
        "benchmark": "measure",
        "indicator": "measure",
        "benchmark/indicator": "measure",
        "benchmark/ indicator": "measure",
        "short title": "short_title",
        "status": "status",
        "updated": "updated",
        "desc body": "body",
        "description body": "body",
        "body": "body",
        "description": "body",
        "other benchmarks": "other",
        "other": "other",
        "other_indicators": "other",
        "influences": "influences",
        "notes": "notes",
        "nsw": "nsw",
        "vic": "vic",
        "qld": "qld",
        "wa": "wa",
        "sa": "sa",
        "tas": "tas",
        "act": "act",
        "nt": "nt",
        "australia": "australia",
        "aust": "australia",
    }
    sheet = wb[sheetname]
    desc = {}
    append_to = None
    row = 1
    if indicator:
        statuses = indicator_statuses
    else:
        statuses = benchmark_statuses
    key_lookup.update(additional_lookups)
    while True:
        try:
            rawkey = sheet["A%d" % row].value
        except IndexError:
            break
        if rawkey:
            key = key_lookup[rawkey.lower()]
        value = sheet["B%d" % row].value
        if not rawkey and not value:
            break
        if key == "status":
            status = None
            for sv in statuses.values():
                if value.lower() in (sv["tlc"], sv["short"].lower()):
                    status = sv
                    break
            if not status:
                raise LoaderException("Unrecognised benchmark status: %s" % value)
            desc["status"] = status
        elif key == "updated":
            desc["updated"] = value
        elif key == "measure":
            desc["measure"] = value
        elif key == "short_title":
            desc["short_title"] = value
        elif key == "body":
            desc["body"] = []
            append_to = desc["body"]
            key = None
        elif key == "influences":
            desc["influences"] = []
            append_to = desc["influences"]
            key = None
        elif key == "notes":
            desc["notes"] = []
            append_to = desc["notes"]
            key = None
        elif key == "other":
            desc["other"] = []
            append_to = desc["other"]
            key = None
        elif key in additional_lookups.keys():
            desc[key_lookup[key]] = []
            append_to = desc[key_lookup[key]]
            key = None
        elif key in ("nsw", "vic", "qld", "wa", "sa", "tas", 
                        "act", "nt", "australia"):
            desc[key] = []
            append_to = desc[key]
            key = None
        if not key and value:
            append_to.append(value)
        row += 1
    return desc

def populate_raw_data(widget_url, label, rds_url,
                    model, field_map, query_kwargs={}, 
                    use_states=True, use_dates=True, pval=None):
    messages = []
    rds = get_rawdataset(widget_url, label, rds_url)
    clear_rawdataset(rds, pval=pval)
    sort_order = 1
    if use_dates:
        order_by_args = [ "state", "year", "financial_year" ]
    else:
        order_by_args = [ "state", ]
    for obj in model.objects.filter(**query_kwargs).order_by(*order_by_args):
        kwargs = {}
        if use_states:
            kwargs["jurisdiction"] = obj.state_display()
        if use_dates:
            kwargs["year"] = obj.year_display()
        for model_field, rds_field in field_map.items():
            if callable(getattr(obj, model_field)):
                kwargs[rds_field] = unicode(getattr(obj, model_field)())
            else:
                kwargs[rds_field] = unicode(getattr(obj, model_field))
        if pval:
            kwargs["pval"] = pval
        add_rawdatarecord(rds, sort_order, **kwargs)
        sort_order += 1
    return messages

def populate_crosstab_raw_data(widget_url, label, rds_url,
                    model, field_map, field_map_states=None, query_kwargs={}, pval=None):
    messages = []
    if field_map_states is None:
        field_map_states = field_map
    rds = get_rawdataset(widget_url, label, rds_url)
    clear_rawdataset(rds, pval=pval)
    sort_order = 1
    kwargs = {}
    kwargs["year"] = None
    if pval:
        kwargs["pval"] = pval
    for obj in model.objects.filter(**query_kwargs).order_by("year", "financial_year", "state"):
        if obj.year_display() != kwargs["year"]:
            if kwargs["year"]:
                add_rawdatarecord(rds, sort_order, **kwargs)
                sort_order += 1
            kwargs = {}
            kwargs["year"] = obj.year_display()
            if pval:
                kwargs["pval"] = pval
        jurisdiction = obj.state_display().lower()
        if obj.state == AUS:
            used_map = field_map
        else:
            used_map = field_map_states
        for model_field, rds_field in used_map.items():
            if callable(getattr(obj, model_field)):
                kwargs[jurisdiction + "_" + rds_field] = unicode(getattr(obj, model_field)())
            else:
                kwargs[jurisdiction + "_" + rds_field] = unicode(getattr(obj, model_field))
    if kwargs["year"]:
        add_rawdatarecord(rds, sort_order, **kwargs)
    return messages

def populate_raw_data_nostate(widget_url, label, rds_url,
                    model, field_map):
    messages = []
    rds = get_rawdataset(widget_url, label, rds_url)
    clear_rawdataset(rds)
    sort_order = 1
    for obj in model.objects.all().filter(state=AUS).order_by("year", "financial_year"):
        kwargs = {
            "year": obj.year_display()
        }
        for model_field, rds_field in field_map.items():
            kwargs[rds_field] = unicode(getattr(obj, model_field))
        add_rawdatarecord(rds, sort_order, **kwargs)
        sort_order += 1
    return messages

def update_graph_data(wurl, wlbl, graphlbl, model, field,
                            jurisdictions = None,
                            benchmark_start=None,
                            benchmark_end=None,
                            benchmark_gen=lambda x: x,
                            use_error_bars=False,
                            verbosity=0,
                            pval=None):
    messages=[]
    g = get_graph(wurl, wlbl, graphlbl)
    clear_graph_data(g, pval=pval)
    if jurisdictions:
        qry = model.objects.filter(state__in=jurisdictions)
    else:
        qry = model.objects.all()
    benchmark_init = None
    benchmark_final = None
    for o in qry:
        value = getattr(o,field)
        if callable(value):
            value = value()
        if o.state == AUS and o.float_year() == benchmark_start:
            benchmark_init = value
            benchmark_final = benchmark_gen(benchmark_init)
        kwargs = {}
        kwargs["horiz_value"] = o.year_as_date()
        if use_error_bars:
            kwargs["val_min"] = value-o.uncertainty
            kwargs["val_max"] = value+o.uncertainty
        if jurisdictions and o.state != AUS:
            dataset_url = "state"
        else:
            dataset_url = o.state_display().lower()
        if pval is not None:
            kwargs["pval"] = pval
        add_graph_data(g, dataset_url,
                value,
                **kwargs)
    if benchmark_init is not None and benchmark_final is not None :
        add_graph_data(g, "benchmark",
                    benchmark_init,
                    horiz_value=float_year_as_date(benchmark_start),
                    pval=pval)
        add_graph_data(g, "benchmark",
                    benchmark_final,
                    horiz_value=float_year_as_date(benchmark_end),
                    pval=pval)
    if verbosity > 2:
        if pval:
            messages.append("Graph %s (%s) updated" % (graphlbl, pval.parameters()["state_abbrev"]))
        else:
            messages.append("Graph %s updated" % graphlbl)
    return messages

txt_block_template = Template("""<div class="coag_description">
    <div class="coag_desc_heading">
    {{ benchmark }}
    </div>
    <div class="coag_desc_body">
        {% for body_elem in desc.body %}
            <p>{{ body_elem|urlize }}</p>
        {% endfor %}
        {% for body_elem in additional_body %}
            <p>{{ body_elem|urlize }}</p>
        {% endfor %}
        {% for inf_elem in desc.influences %}
            <p>
                {% if forloop.first %}
                    <b>Influences:</b>
                {% endif %}
                {{ inf_elem|urlize }}
            </p>
        {% endfor %}
        {% for other_elem in desc.other %}
            <p>
                {% if forloop.first %}
                    <b>Other Benchmarks/Indicators:</b>
                {% endif %}
                {{ other_elem }}
            </p>
        {% endfor %}
        {% if state_content %}
            {% for sc_elem in state_content %}
                <p>{{ sc_elem|urlize }}</p>
            {% endfor %}
        {% endif %}
    </div>
    {% for note in desc.notes %}
        {% if forloop.first %}
            <div class="coag_desc_notes">
                <p>Notes:</p>
                <ol>
        {% endif %}
                    <li>{{ note|urlize }}</li>
        {% if forloop.last %}
                </ol>
            </div>
        {% endif %}
    {% endfor %}
</div>""")

def update_stats(desc, benchmark, 
                wurl_hero, wlbl_hero, 
                wurl_hero_state, wlbl_hero_state,
                wurl, wlbl, 
                wurl_state, wlbl_state,
                verbosity=0,
                additional_desc_body="additional_body"):
    messages = []
    if wurl_hero:
        set_statistic_data(wurl_hero, wlbl_hero,
                    "status_header",
                    desc["status"]["short"],
                    traffic_light_code=desc["status"]["tlc"],
                    icon_code=desc["status"]["icon"])
        set_actual_frequency_display_text(wurl_hero, wlbl_hero,
                    "%s" % unicode(desc["updated"]))
    if wurl_hero_state:
        p = Parametisation.objects.get(url="state_param")
        for pval in p.parametisationvalue_set.all():
            set_statistic_data(wurl_hero_state, wlbl_hero_state,
                        "status_header",
                        "Australia - " + desc["status"]["short"],
                        traffic_light_code=desc["status"]["tlc"],
                        icon_code=desc["status"]["icon"],
                        pval=pval)
            set_actual_frequency_display_text(wurl_hero_state, wlbl_hero_state,
                        "%s" % unicode(desc["updated"]),
                        pval=pval)
    if wurl:
        set_statistic_data(wurl, wlbl,
                        "status_header",
                        desc["status"]["short"],
                        traffic_light_code=desc["status"]["tlc"],
                        icon_code=desc["status"]["icon"])
        set_statistic_data(wurl, wlbl,
                        "status_short",
                        desc["status"]["short"],
                        traffic_light_code=desc["status"]["tlc"],
                        icon_code=desc["status"]["icon"])
        set_statistic_data(wurl, wlbl,
                        "status_long",
                        desc["status"]["long"],
                        traffic_light_code=desc["status"]["tlc"])
        set_actual_frequency_display_text(wurl, wlbl,
                    "Updated: %s" % unicode(desc["updated"]))
        set_text_block(wurl, wlbl,
                    txt_block_template.render(Context({ 
                                    "benchmark": desc.get("measure", benchmark), 
                                    "additional_body": desc.get(additional_desc_body),
                                    "desc": desc,
                                    "state": "Australia",
                                    "state_content": desc.get("australia") })))
    if wurl_state:
        p = Parametisation.objects.get(url="state_param")
        for pval in p.parametisationvalue_set.all():
            set_statistic_data(wurl_state, wlbl_state,
                            "status_header",
                            "Australia - " + desc["status"]["short"],
                            traffic_light_code=desc["status"]["tlc"],
                            icon_code=desc["status"]["icon"],
                            pval=pval)
            set_statistic_data(wurl_state, wlbl_state,
                            "status_short",
                            "Australia - " + desc["status"]["short"],
                            traffic_light_code=desc["status"]["tlc"],
                            icon_code=desc["status"]["icon"],
                            pval=pval)
            set_statistic_data(wurl_state, wlbl_state,
                            "status_long",
                            desc["status"]["long"],
                            traffic_light_code=desc["status"]["tlc"],
                            pval=pval)
            set_actual_frequency_display_text(wurl_state, wlbl_state,
                        "Updated: %s" % unicode(desc["updated"]),
                        pval=pval)
            st_abbrev = pval.parameters()["state_abbrev"]
            set_text_block(wurl_state, wlbl_state,
                        txt_block_template.render(Context({ 
                                        "benchmark": desc.get("measure", benchmark), 
                                        "additional_body": desc.get(additional_desc_body),
                                        "desc": desc,
                                        "state": st_abbrev,
                                        "state_content": desc.get(st_abbrev.lower()) })),
                        pval=pval)
    if verbosity > 1:
        messages.append("Stats updated")
    return messages

def update_state_stats(wurl_hero, wlbl_hero, wurl_dtl, wlbl_dtl,
                    model, fields,
                    query_filter_kwargs={},
                    want_increase=True,
                    combine_all=False,
                    override_status=None,
                    restrict_states=None,
                    use_benchmark_tls=False,
                    no_data_override=None,
                    status_func=None,
                    status_func_kwargs={},
                    verbosity=0):
    messages = []
    try:
        if len(fields) != len(want_increase):
            raise LoaderError("fields/want_increase array length mismatch")
    except TypeError:
        want_increase = [ want_increase ] * len(fields)
    if use_benchmark_tls:
        statuses = benchmark_statuses
    else:
        statuses = indicator_statuses
    p = Parametisation.objects.get(url="state_param")
    for pval in p.parametisationvalue_set.all():
        state_abbrev = pval.parameters()["state_abbrev"]
        state_num = state_map[state_abbrev]
        if restrict_states and state_num not in restrict_states:
            continue
        if override_status:
            status = statuses[override_status]
        else:
            qry = model.objects.filter(state=state_num, **query_filter_kwargs)
            reference = qry.first()
            measure = qry.last()
            if reference is None:
                if verbosity > 1:
                    messages.append("%s: No data" % state_abbrev)
                status = indicator_statuses["no_data"]
            elif reference == measure:
                status = indicator_statuses["new_indicator"]
                if verbosity > 1:
                    messages.append("%s: New indicator" % state_abbrev)
            elif status_func:
                if combine_all:
                    my_statuses = []
                    for obj in qry:
                        status = status_func(obj, **status_func_kwargs)
                        if status not in my_statuses:
                           my_statuses.append(status)
                    if len(my_statuses) == 0:
                        if no_data_override:
                           status = no_data_override
                        elif use_benchmark_tls:
                            status = "new_benchmark"
                        else:
                            status = "no_data"
                    elif len(my_statuses) == 1:
                        status = my_statuses[0]
                    else:
                        status = "mixed_results"
                    status = statuses[status]
                else:
                    status = statuses[status_func(measure, **status_func_kwargs)]
            else:
                my_statuses = []
                for flds, want_incr in zip(fields, want_increase):
                    field = flds[0]
                    uncertainty_field = flds[1]
                    if len(flds) > 2:
                        rse_field = flds[2]
                    else:
                        rse_field = None
                    val_1 = getattr(reference, field)
                    val_2 = getattr(measure, field)
                    if callable(val_1):
                        val_1 = val_1()
                    if callable(val_2):
                        val_2 = val_2()
                    if rse_field:
                        rse_1 = getattr(reference, rse_field)
                        rse_2 = getattr(measure, rse_field)
                    else:
                        rse_1 = None
                        rse_2 = None
                    if uncertainty_field:
                        uncertainty_1 = getattr(reference, uncertainty_field)
                        uncertainty_2 = getattr(measure, uncertainty_field)
                    else:
                        uncertainty_1 = None
                        uncertainty_2 = None
                    my_statuses.append(indicator_status_tlc(
                                                val_1, val_2, 
                                                uncertainty_1, uncertainty_2, 
                                                rse_1, rse_2, 
                                                want_incr))
                st_so_far = None
                for st in my_statuses:
                    if not st_so_far:
                        st_so_far = st
                    elif st_so_far != st:
                        st_so_far = "mixed_results"
                        break
                status = indicator_statuses[st_so_far]
        if wurl_hero:
            set_statistic_data(wurl_hero, wlbl_hero,
                            "status_header_state",
                            state_abbrev + " - " + status["short"],
                            traffic_light_code=status["tlc"],
                            icon_code=status["icon"],
                            pval=pval)
        if wurl_dtl:
            set_statistic_data(wurl_dtl, wlbl_dtl,
                        "status_short_state",
                        state_abbrev + " - " + status["short"],
                        traffic_light_code=status["tlc"],
                        icon_code=status["icon"],
                        pval=pval)
            set_statistic_data(wurl_dtl, wlbl_dtl,
                        "status_header_state",
                        state_abbrev + " - " + status["short"],
                        traffic_light_code=status["tlc"],
                        icon_code=status["icon"],
                        pval=pval)
    return messages

def indicator_status_tlc(val_1, val_2, uncertainty_1=None, uncertainty_2=None, err_1=None, err_2=None, want_incr=True):
    diff = val_2 - val_1
    diff = val_2 - val_1
    if isinstance(diff, float):
        diff = Decimal(diff).quantize(Decimal('0.00001'), rounding=ROUND_HALF_UP)
    if isinstance(diff, int):
        diff = Decimal(diff)
    val_1 = float(val_1) / 100.0
    val_2 = float(val_2) / 100.0
    if err_1:
        err_1 = float(err_1) / 100.0
    if err_2:
        err_2 = float(err_2) / 100.0
    if err_1 and err_2:
        significance = math.sqrt(
                    math.pow(val_1 * err_1, 2) 
                    + math.pow(val_2 * err_2, 2)
        )
        test_stat = float(diff)/100.0/significance
        significant = abs(test_stat) > 1.96
        if not significant:
            return "no_improvement"
        elif (want_incr and diff == abs(diff)) or (not want_incr and diff != abs(diff)):
            return "improving"
        else:
            return "negative_change"
    if uncertainty_1 is not None and uncertainty_2 is not None:
        total_uncertainty = uncertainty_1 + uncertainty_2
    else:
        total_uncertainty = Decimal("0.0")
    if abs(diff) < total_uncertainty or diff.is_zero():
        return "no_improvement"
    elif (want_incr and diff == abs(diff)) or (not want_incr and diff != abs(diff)):
        return "improving"
    else:
        return "negative_change"

def indicator_tlc_trend(ref_val, val, ref_uncertainty=None, uncertainty=None, ref_rse=None, rse=None, want_increase=True):
    direction = cmp(val, ref_val)
    return (indicator_status_tlc(ref_val, val, ref_uncertainty, uncertainty, ref_rse, rse, want_increase), direction)

