#   Copyright 2016 CSIRO
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.


import datetime
import csv
from decimal import Decimal, ROUND_HALF_UP
import re
from openpyxl import load_workbook
from django.db.models import Sum
from dashboard_loader.loader_utils import *
from coag_uploader.models import *
from infrastructure_npa_uploader.models import *
from coag_uploader.uploader import load_state_grid, load_benchmark_description, update_graph_data, populate_raw_data, populate_crosstab_raw_data, update_stats, update_state_stats, column_labels
from widget_def.models import Parametisation



# These are the names of the groups that have permission to upload data for this uploader.
# If the groups do not exist they are created on registration.
groups = [ "upload_all", "upload" ]

# This describes the file format.  It is used to describe the format by both
# "python manage.py upload_data frontlineservice_uploader" and by the uploader 
# page in the data admin GUI.


file_format = {
    "format": "xlsx",
    "sheets": [
            {
                "name": "Projects",
                "cols": [ 
                            ('A', 'Year e.g. 2007-08 or 2007/08 or 2007'),
                            ('B', 'Status ("Complete", "Underway", "Pending" or "Total")'),
                            ('C', 'Metric ("Project Numbers", "Total Project Cost")'),
                            ('...', 'Column per state'),
                        ],
                "rows": [
                            ('1', "Heading row"),
                            ('2', "State Heading row"),
                            ('...', 'Eight rows of per year, one for each combination of status and metric.  Note that rows with status "Total" are ignored - all displayed totals are recalculated',)
                        ],
                "notes": [
                    'Blank rows and columns ignored',
                    'Aust column represents cross-jurisdictional projects',
                ],
            },
            {
                "name": "Milestones",
                "cols": [ 
                            ('A', 'State'),
                            ('B', 'Key Project")'),
                            ('C', 'Project Milestones'),
                            ('D', 'Project Status'),
                        ],
                "rows": [
                            ('1', "Heading row"),
                            ('2', "Column Heading row"),
                            ('...', 'Row per key project'),
                        ],
                "notes": [
                    'Blank rows and columns ignored',
                ],
            },
            {
                "name": "Description",
                "cols": [
                            ('A', 'Key'),
                            ('B', 'Value'),
                        ],
                "rows": [
                            ('Measure', 'Full description of benchmark'),
                            ('Short Title', 'Short widget title (not used)'),
                            ('Status', 'Benchmark status'),
                            ('Updated', 'Year data last updated'),
                            ('Desc body', 'Body of benchmark status description. One paragraph per line.'),
                            ('Influences', '"Influences" text of benchmark status description. One paragraph per line (optional)'),
                            ('Notes', 'Notes for benchmark status description.  One note per line.'),
                        ],
                "notes": [
                         ],
            }
        ],
}

def upload_file(uploader, fh, actual_freq_display=None, verbosity=0):
        messages = []
#try:
        if verbosity > 0:
            messages.append("Loading workbook...")
        wb = load_workbook(fh, read_only=True)
        InfrastructureProjects.objects.all().delete()
        messages.extend(
                load_state_grid(wb, "Projects",
                                "Infrastructure", "Projects",
                                None, InfrastructureProjects,
                                {}, {
                                        "completed": ("Completed", "Project Numbers"),
                                        "underway": ("Underway", "Project Numbers"),
                                        "pending": ("Pending", "Project Numbers"),
                                        "completed_cost": ("Completed", "Total Project Cost"),
                                        "underway_cost": ("Underway", "Total Project Cost"),
                                        "pending_cost": ("Pending", "Total Project Cost"),
                                    },
                                verbosity=verbosity)
                )
        messages.extend(
                load_progress_grid(wb, "Milestones",
                                "Infrastructure", "Key Project Milestones",
                                InfrastructureKeyProjects,
                                {
                                    "project": "Key Project",
                                    "milestones": "Milestones",
                                    "state": "State",
                                    "status": "Status",
                                },
                                transforms = {
                                    "state": lambda x: state_map[x],
                                    "status": lambda x:project_map[x],
                                },
                                verbosity=verbosity)
        )
        desc = load_benchmark_description(wb, "Description")
        messages.extend(update_stats(desc, None,
                            "projects-infrastructure-hero", "projects-infrastructure-hero", 
                            "projects-infrastructure-hero-state", "projects-infrastructure-hero-state", 
                            "infrastructure_projects", "infrastructure_projects", 
                            "infrastructure_projects_state", "infrastructure_projects_state", 
                            verbosity))
        messages.extend(update_state_stats(
                            "projects-infrastructure-hero-state", "projects-infrastructure-hero-state", 
                            "infrastructure_projects_state", "infrastructure_projects_state", 
                            None, [],
                            override_status=desc["status"]["tlc"],
                            use_benchmark_tls = True,
                            verbosity=verbosity))
        aust_sums = InfrastructureProjects.objects.aggregate(Sum('completed'), Sum('underway'), Sum('pending'))
        set_statistic_data("projects-infrastructure-hero", "projects-infrastructure-hero",
                            "pending", float(aust_sums["pending__sum"]),
                            traffic_light_code=desc["status"]["tlc"])
        set_statistic_data("projects-infrastructure-hero", "projects-infrastructure-hero",
                            "completed", float(aust_sums["completed__sum"]),
                            traffic_light_code=desc["status"]["tlc"])
        set_statistic_data("projects-infrastructure-hero", "projects-infrastructure-hero",
                            "underway", float(aust_sums["underway__sum"]),
                            traffic_light_code=desc["status"]["tlc"])
        set_statistic_data("infrastructure_projects", "infrastructure_projects", 
                            "pending", float(aust_sums["pending__sum"]),
                            traffic_light_code=desc["status"]["tlc"])
        set_statistic_data("infrastructure_projects", "infrastructure_projects", 
                            "completed", float(aust_sums["completed__sum"]),
                            traffic_light_code=desc["status"]["tlc"])
        set_statistic_data("infrastructure_projects", "infrastructure_projects", 
                            "underway", float(aust_sums["underway__sum"]),
                            traffic_light_code=desc["status"]["tlc"])
        clear_statistic_list( "infrastructure_projects", "infrastructure_projects", 
                              "key_projects")
        sort_order = 5
        for kp in InfrastructureKeyProjects.objects.all():
            add_statistic_list_item(
                            "infrastructure_projects", "infrastructure_projects", 
                            "key_projects",
                            kp.status_display(),
                            sort_order,
                            traffic_light_code = kp.status_tlc(),
                            label="(%s) %s" % (kp.state_display(), kp.project)
            )
            sort_order += 5
        messages.extend(
                populate_raw_data(
                            "infrastructure_projects", "infrastructure_projects", 
                            "infrastructure_projects", InfrastructureProjects, 
                                {
                                    "completed": "projects_completed",
                                    "underway": "projects_underway",
                                    "pending": "projects_pending",
                                    "total": "projects_total",
                                    "completed_cost": "cost_completed",
                                    "underway_cost": "cost_underway",
                                    "pending_cost": "cost_pending",
                                    "total_cost": "cost_total",
                                },
                            use_dates=False)
                )
        messages.extend(
                populate_raw_data(
                            "infrastructure_projects", "infrastructure_projects", 
                            "infrastructure_projects_status", InfrastructureKeyProjects, 
                                {
                                    "project": "project",
                                    "milestones": "milestones",
                                    "status_display": "status",
                                },
                            use_dates=False)
                )
        messages.extend(
                populate_raw_data(
                            "infrastructure_projects", "infrastructure_projects", 
                                "data_table", InfrastructureProjects, 
                                {
                                    "completed": "complete_count",
                                    "underway": "underway_count",
                                    "pending": "pending_count",
                                    "total": "total_count",
                                    "completed_cost": "complete_cost",
                                    "underway_cost": "underway_cost",
                                    "pending_cost": "pending_cost",
                                    "total_cost": "total_cost",
                                },
                                use_dates=False)
                )
        p = Parametisation.objects.get(url="state_param")
        for pval in p.parametisationvalue_set.all():
            state_num = state_map[pval.parameters()["state_abbrev"]]
            state_counts = InfrastructureProjects.objects.get(state=state_num)
            set_statistic_data(
                        "projects-infrastructure-hero-state", "projects-infrastructure-hero-state",
                        "completed", float(aust_sums["completed__sum"]),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "projects-infrastructure-hero-state", "projects-infrastructure-hero-state",
                        "underway", float(aust_sums["underway__sum"]),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "projects-infrastructure-hero-state", "projects-infrastructure-hero-state",
                        "pending", float(aust_sums["pending__sum"]),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "projects-infrastructure-hero-state", "projects-infrastructure-hero-state",
                        "completed_state", float(state_counts.completed),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "projects-infrastructure-hero-state", "projects-infrastructure-hero-state",
                        "underway_state", float(state_counts.underway),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "projects-infrastructure-hero-state", "projects-infrastructure-hero-state",
                        "pending_state", float(state_counts.pending),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "infrastructure_projects_state", "infrastructure_projects_state", 
                        "completed", float(aust_sums["completed__sum"]),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "infrastructure_projects_state", "infrastructure_projects_state", 
                        "underway", float(aust_sums["underway__sum"]),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "infrastructure_projects_state", "infrastructure_projects_state", 
                        "pending", float(aust_sums["pending__sum"]),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "infrastructure_projects_state", "infrastructure_projects_state", 
                        "completed_state", float(state_counts.completed),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "infrastructure_projects_state", "infrastructure_projects_state", 
                        "underway_state", float(state_counts.underway),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            set_statistic_data(
                        "infrastructure_projects_state", "infrastructure_projects_state", 
                        "pending_state", float(state_counts.pending),
                        traffic_light_code=desc["status"]["tlc"],
                        pval=pval)
            clear_statistic_list( "infrastructure_projects_state", "infrastructure_projects_state", 
                                "key_projects",
                              pval=pval)
            sort_order = 5
            for kp in InfrastructureKeyProjects.objects.filter(state=state_num):
                add_statistic_list_item(
                                "infrastructure_projects_state", "infrastructure_projects_state", 
                                "key_projects",
                                kp.status_display(),
                                sort_order,
                                traffic_light_code = kp.status_tlc(),
                                label=kp.project,
                                pval=pval
                )
                sort_order += 5
            messages.extend(
                    populate_raw_data(
                                "infrastructure_projects_state", "infrastructure_projects_state", 
                                "infrastructure_projects", InfrastructureProjects, 
                                    {
                                        "completed": "projects_completed",
                                        "underway": "projects_underway",
                                        "pending": "projects_pending",
                                        "total": "projects_total",
                                        "completed_cost": "cost_completed",
                                        "underway_cost": "cost_underway",
                                        "pending_cost": "cost_pending",
                                        "total_cost": "cost_total",
                                    },
                                use_dates=False,
                                pval=pval)
                    )
            messages.extend(
                    populate_raw_data(
                                "infrastructure_projects_state", "infrastructure_projects_state", 
                                "infrastructure_projects_status", InfrastructureKeyProjects, 
                                    {
                                        "project": "project",
                                        "milestones": "milestones",
                                        "status_display": "status",
                                    },
                                use_dates=False,
                                pval=pval)
                    )
            messages.extend(
                    populate_raw_data(
                                "infrastructure_projects_state", "infrastructure_projects_state", 
                                    "data_table", InfrastructureProjects, 
                                    {
                                        "completed": "complete_count",
                                        "underway": "underway_count",
                                        "pending": "pending_count",
                                        "total": "total_count",
                                        "completed_cost": "complete_cost",
                                        "underway_cost": "underway_cost",
                                        "pending_cost": "pending_cost",
                                        "total_cost": "total_cost",
                                    },
                                    use_dates=False,
                                    pval=pval)
                    )
#    except LoaderException, e:
#        raise e
#    except Exception, e:
#        raise LoaderException("Invalid file: %s" % unicode(e))
        return messages

def load_progress_grid(wb, sheet_name, data_category, dataset, model,
                                cell_rows, transforms={}, verbosity=0):
    messages = []
    if verbosity > 2:
        messages.append("Loading %s Data: %s" % (data_category, dataset))
    sheet = wb[sheet_name]
    row = 1
    records_written = 0
    column_map = {}
    columns_mapped = False
    model.objects.all().delete()
    while True:
        if not columns_mapped:
            for col in column_labels(1, sheet.max_column):
                try:
                    cval = sheet["%s%d" % (col, row)].value
                    if cval is not None:
                        cval = cval.strip()
                except IndexError:
                    raise LoaderException("Sheet finished, column headings not found")
                for k, v in cell_rows.items():
                    if cval == v:
                        column_map[k] = col
            columns_mapped = True
            for col in cell_rows.keys():
                if col not in column_map:
                    columns_mapped = False
                    break
        else:
            kwargs = {}
            for fldname, col in column_map.items():
                try:
                    cval = sheet["%s%d" % (col, row)].value
                except IndexError:
                    if verbosity > 1:
                        messages.append("Records written: %d" % records_written)
                    return messages
                if cval is None:
                    break
                cval = cval.strip()
                if fldname in transforms:
                    cval = transforms[fldname](cval)
                kwargs[fldname] = cval
            if kwargs:
                obj = model(**kwargs)
                obj.save()
                records_written += 1
        row += 1
