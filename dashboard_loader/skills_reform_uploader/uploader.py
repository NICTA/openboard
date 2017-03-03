#   Copyright 2017 CSIRO
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.


import datetime
import csv
from decimal import Decimal, ROUND_HALF_UP
import re
from openpyxl import load_workbook
from django.db.models import Sum
from dashboard_loader.loader_utils import *
from coag_uploader.models import *
from skills_reform_uploader.models import *
from coag_uploader.uploader import load_state_grid, load_benchmark_description, update_graph_data, populate_raw_data, populate_crosstab_raw_data, update_stats, update_state_stats, column_labels
from widget_def.models import Parametisation



# These are the names of the groups that have permission to upload data for this uploader.
# If the groups do not exist they are created on registration.
groups = [ "upload_all", "upload" ]

# This describes the file format.  It is used to describe the format by both
# "python manage.py upload_data frontlineservice_uploader" and by the uploader 
# page in the data admin GUI.


file_format = {
    "format": "xlsx",
    "sheets": [
            {
                "name": "Data",
                "cols": [ 
                            ('A', 'Assessment'),
                            ('B', 'Performance Indicator'),
                            ('C', 'Metric ("Project Numbers", "Total Project Cost")'),
                            ('...', 'Column per state and Commonwealth("Aust")'),
                        ],
                "rows": [
                            ('1', "Heading row"),
                            ('2', "State Heading row"),
                            ('...', 'One row per performance indicator, containing status for each jurisdiction',)
                        ],
                "notes": [
                    'Blank rows and columns ignored',
                ],
            },
            {
                "name": "Description",
                "cols": [
                            ('A', 'Key'),
                            ('B', 'Value'),
                        ],
                "rows": [
                            ('Measure', 'Full description of benchmark'),
                            ('Short Title', 'Short widget title (not used)'),
                            ('Status', 'Benchmark status'),
                            ('Updated', 'Year data last updated'),
                        ],
                "notes": [
                         ],
            }
        ],
}

def upload_file(uploader, fh, actual_freq_display=None, verbosity=0):
    messages = []
    try:
        if verbosity > 0:
            messages.append("Loading workbook...")
        wb = load_workbook(fh, read_only=True)
        messages.extend(
                load_skills_reform_progress_grid(wb, "Data", verbosity=verbosity)
        )
        desc = load_benchmark_description(wb, "Description")
        messages.extend(update_stats(desc, None,
                            "skills_reform-skills-hero", "skills_reform-skills-hero", 
                            "skills_reform-skills-hero-state", "skills_reform-skills-hero-state", 
                            "skills_reform", "skills_reform", 
                            "skills_reform_state", "skills_reform_state", 
                            verbosity))
        messages.extend(update_state_stats(
                            "skills_reform-skills-hero-state", "skills_reform-skills-hero-state", 
                            "skills_reform_state", "skills_reform_state", 
                            None, [],
                            override_status=desc["status"]["tlc"],
                            use_benchmark_tls = True,
                            verbosity=verbosity))
        aust_completed = SkillsReformData.objects.filter(state=AUS).filter(status="Completed").count()
        set_statistic_data("skills_reform-skills-hero", "skills_reform-skills-hero", 
                            "completed", aust_completed,
                            traffic_light_code=desc["status"]["tlc"])
        set_statistic_data("skills_reform", "skills_reform", 
                            "completed", aust_completed,
                            traffic_light_code=desc["status"]["tlc"])
        messages.extend(populate_my_raw_dataset("skills_reform", "skills_reform", "skills_reform"))
        messages.extend(populate_my_raw_dataset("skills_reform", "skills_reform", "data_table"))
        p = Parametisation.objects.get(url="state_param")
        for pval in p.parametisationvalue_set.all():
            state_num = state_map[pval.parameters()["state_abbrev"]]
            state_completed = SkillsReformData.objects.filter(state=state_num).filter(status="Completed").count()
            set_statistic_data("skills_reform-skills-hero-state", "skills_reform-skills-hero-state", 
                                "completed", state_completed,
                                traffic_light_code=desc["status"]["tlc"],
                                pval=pval)
            set_statistic_data("skills_reform_state", "skills_reform_state", 
                                "completed", state_completed,
                                traffic_light_code=desc["status"]["tlc"],
                                pval=pval)
            messages.extend(populate_my_raw_dataset("skills_reform_state", "skills_reform_state", 
                                    "skills_reform", pval=pval))
            messages.extend(populate_my_raw_dataset("skills_reform_state", "skills_reform_state", 
                                    "data_table", pval=pval))
    except LoaderException, e:
        raise e
    except Exception, e:
        raise LoaderException("Invalid file: %s" % unicode(e))
    return messages

def load_skills_reform_progress_grid(wb, sheet_name, verbosity=0):
    messages = []
    if verbosity > 2:
        messages.append("Loading Skills Reform Data")
    sheet = wb[sheet_name]
    row = 1
    records_written = 0
    column_map = {}
    columns_mapped = False
    SkillsReformData.objects.all().delete()
    expected_columns = {
            "assessment": "Assessment",
            "perf_ind": "Performance Indicator",
            "nsw": "NSW",
            "vic": "Vic",
            "qld": "Qld",
            "wa": "WA",
            "sa": "SA",
            "tas": "Tas",
            "act": "ACT",
            "nt": "NT",
            "australia": "Aust"
    }
    state_cols = [
        "nsw", "vic", "qld",
        "wa", "sa", "tas",
        "act", "nt", "australia",
        ]
    last_assessment = None
    assessment_sort_order = 0
    perfind_sort_order = 0
    blank_rows = 0
    while True:
        if not columns_mapped:
            for col in column_labels(1, sheet.max_column):
                try:
                    cval = sheet["%s%d" % (col, row)].value
                    if cval is not None:
                        cval = cval.strip()
                except IndexError:
                    raise LoaderException("Sheet finished, column headings not found")
                if cval is None:
                    continue
                for k, v in expected_columns.items():
                    if cval.lower() == v.lower():
                        column_map[k] = col
            columns_mapped = True
            for expcol in expected_columns.keys():
                if expcol not in column_map:
                    columns_mapped = False
                    break
        else:
            kwargs = {}
            try:
                cval = sheet["%s%d" % (column_map["assessment"], row)].value
            except IndexError:
                if verbosity > 1:
                    messages.append("Records written: %d" % records_written)
                return messages
            if cval is None:
                blank_rows += 1
                if blank_rows > 100:
                    if verbosity > 1:
                        messages.append("Records written: %d" % records_written)
                    return messages
                continue
            cval=cval.strip()
            kwargs["assessment"]=cval
            if cval != last_assessment:
                last_assessment = cval
                assessment_sort_order += 10
                perfind_sort_order = 0
            kwargs["assessment_sort_order"]=assessment_sort_order
            perfind_sort_order += 10
            cval = sheet["%s%d" % (column_map["perf_ind"], row)].value
            cval=cval.strip()
            kwargs["performance_indicator"]=cval
            kwargs["performance_indicator_sort_order"]=perfind_sort_order
            for juris in state_cols:
                cval = sheet["%s%d" % (column_map[juris], row)].value
                cval = cval.strip()
                kwargs["state"] = state_map[juris]
                kwargs["status"] = cval
                obj = SkillsReformData(**kwargs)
                obj.save()
                records_written += 1
        row += 1

def populate_my_raw_dataset(wurl, wlbl, rdsname, pval=None):
    messages = []
    rds = get_rawdataset(wurl, wlbl, rdsname)
    clear_rawdataset(rds, pval=pval)
    sort_order = 1
    last_pi = None
    kwargs = {}
    for obj in SkillsReformData.objects.all().order_by("assessment_sort_order", "performance_indicator_sort_order", "state"):
        if last_pi is None or last_pi!=obj.performance_indicator:
            if last_pi is not None:
                add_rawdatarecord(rds, sort_order, pval=pval, **kwargs)
                sort_order += 1
            last_pi = obj.performance_indicator
            kwargs = {}
            kwargs["assessment"] = obj.assessment
            kwargs["performance_indicator"] = obj.performance_indicator
        kwargs[obj.state_display().lower() + "_status"] = obj.status
    return messages

