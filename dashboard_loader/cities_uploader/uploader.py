#   Copyright 2016 NICTA
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.

import csv
import decimal
import re
from openpyxl import load_workbook
from dashboard_loader.loader_utils import *
from widget_def.models import Parametisation, ParametisationValue

# These are the names of the groups that have permission to upload data for this uploader.
# If the groups do not exist they are created on registration.
groups = [ "upload_all", "upload_cities" ]

# This describes the file format.  It is used to describe the format by both
# "python manage.py upload_data frontlineservice_uploader" and by the uploader 
# page in the data admin GUI.
file_format = {
    # format:  Either "csv", "xls", "xlsx" or "zip"
    "format": "xlsx",
    # sheets: csv formats should only have one sheet.  For "zip", sheets are 
    #         csv files within the zip file
    "sheets": [
        {
            "name": "Example design",
            "cols": [],
            "rows": [],
            "notes": ["For human use only - not read by uploader"],
        },
        {
            "name": "SA4",
            "cols": [
                    ('A', 'SA4 Code (used to identify SA4 region)'),
                    ('B', 'SA4 Name (ignored)'),

                    ('C', 'Life Expectancy at birth (years) (reference year)'),
                    ('D', 'Life Expectancy at birth (years) (current year)'),
                    ('E', 'Life Expectancy at birth (percentile) (reference year)'),
                    ('F', 'Life Expectancy at birth (percentile) (current year)'),

                    ('G', 'Adults who are overweight or obese (percent) (reference year)'),
                    ('H', 'Adults who are overweight or obese (percent) (current year)'),
                    ('I', 'Adults who are overweight or obese (percentile) (reference year)'),
                    ('J', 'Adults who are overweight or obese (percentile) (current year)'),

                    ('K', 'Homeless persons per 10k population (reference year)'),
                    ('L', 'Homeless persons per 10k population (current year)'),
                    ('M', 'Homeless persons per 10k population (percentile) (reference year)'),
                    ('N', 'Homeless persons per 10k population (percentile) (current year)'),

                    ('O', 'Proportion of households that own their own home (percentage) (reference year)'),
                    ('P', 'Proportion of households that own their own home (percentage) (current year)'),
                    ('Q', 'Proportion of households that own their own home (percentile) (reference year)'),
                    ('R', 'Proportion of households that own their own home (percentile) (current year)'),

                    ('S', 'People with vocational or higher education qualification (percentage) (reference year)'),
                    ('T', 'People with vocational or higher education qualification (percentage) (current year)'),
                    ('U', 'People with vocational or higher education qualification (percentile) (reference year)'),
                    ('V', 'People with vocational or higher education qualification (percentile) (current year)'),

                    ('W', 'State Code'),

                    ('X', 'Life Expectancy at birth (State) (years) (reference year)'),
                    ('Y', 'Life Expectancy at birth (State) (years) (current year)'),
                    
                    ('Z', 'Adults who are overweight or obese (State) (percent) (reference year)'),
                    ('AA', 'Adults who are overweight or obese (State) (percent) (current year)'),
                    ('AB', 'Homeless persons per 10k population (State) (reference year)'),
                    ('AC', 'Homeless persons per 10k population (State) (current year)'),
                    ('AD', 'Proportion of households that own their own home (State) (percentage) (reference year)'),
                    ('AE', 'Proportion of households that own their own home (State) (percentage) (current year)'),
                    ('AF', 'People with vocational or higher education qualification (State) (percentage) (reference year)'),
                    ('AG', 'People with vocational or higher education qualification (State) (percentage) (current year)'),

                    ('AH', 'Labour force participation (percent) (reference year)'),
                    ('AI', 'Labour force participation (percent) (current year)'),
                    ('AJ', 'Labour force participation (% rank) (reference year)'),
                    ('AK', 'Labour force participation (% rank) (current year)'),

                    ('AL', 'Proportion who are earning or learning (percent) (reference year)'),
                    ('AM', 'Proportion who are earning or learning (percent) (current year)'),
                    ('AN', 'Proportion who are earning or learning (% rank) (reference year)'),
                    ('AO', 'Proportion who are earning or learning (% rank) (current year)'),

                    ('AP', 'Real median weekly household income ($) (reference year)'),
                    ('AQ', 'Real median weekly household income ($) (current year)'),
                    ('AR', 'Real median weekly household income (% rank) (reference year)'),
                    ('AS', 'Real median weekly household income (% rank) (current year)'),

                    ('AT', 'Ratio (P10:P50) (reference year)'),
                    ('AU', 'Ratio (P10:P50) (current year)'),
                    ('AV', 'Ratio (P10:P50) (% rank) (reference year)'),
                    ('AW', 'Ratio (P10:P50) (% rank) (current year)'),
                    
                    ('AX', 'New business entry rate (percent) (reference year)'),
                    ('AY', 'New business entry rate (percent) (current year)'),
                    ('AZ', 'New business entry rate (% rank) (reference year)'),
                    ('BA', 'New business entry rate (% rank) (current year)'),

                    ('BB', 'Labour force participation (State) (percent) (reference year)'),
                    ('BC', 'Labour force participation (State) (percent) (current year)'),
                    ('BD', 'Proportion who are earning or learning (State) (percent) (reference year)'),
                    ('BE', 'Proportion who are earning or learning (State) (percent) (current year)'),
                    ('BF', 'Real median weekly household income ($) (State) (reference year)'),
                    ('BG', 'Real median weekly household income ($) (State) (current year)'),
                    ('BH', 'Ratio (P10:P50) (State) (reference year)'),
                    ('BI', 'Ratio (P10:P50) (State) (current year)'),
                    ('BJ', 'New business entry rate (State) (percent) (reference year)'),
                    ('BK', 'New business entry rate (State) (percent) (current year)'),

                    ('BL', 'Terrestrial area protected (percent) (reference year)'),
                    ('BM', 'Terrestrial area protected (percent) (current year)'),
                    ('BN', 'Terrestrial area protected (rank) (reference year)'),
                    ('BO', 'Terrestrial area protected (rank) (current year)'),
        
                    ('BP', 'Domestic involving nature activity trip (visits per resident) (reference year)'),
                    ('BQ', 'Domestic involving nature activity trip (visits per resident) (current year)'),
                    ('BR', 'Domestic involving nature activity trip (rank) (reference year)'),
                    ('BS', 'Domestic involving nature activity trip (rank) (current year)'),
                    
                    ('BT', 'Terrestrial area protected (State) (percent) (reference year)'),
                    ('BU', 'Terrestrial area protected (State) (percent) (current year)'),
                    ('BV', 'Domestic involving nature activity trip (State) (visits per resident) (reference year)'),
                    ('BW', 'Domestic involving nature activity trip (State) (visits per resident) (current year)'),

                    ('BX', 'Overseas-born Australian residents who are Australian citizens (percent) (reference year)'),
                    ('BY', 'Overseas-born Australian residents who are Australian citizens (percent) (current year)'),
                    ('BZ', 'Overseas-born Australian residents who are Australian citizens (rank) (reference year)'),
                    ('CA', 'Overseas-born Australian residents who are Australian citizens (rank) (current year)'),
                    ('CB', 'Overseas-born Australian residents who are Australian citizens (State) (percent) (reference year)'),
                    ('CC', 'Overseas-born Australian residents who are Australian citizens (State) (percent) (current year)'),
        
                    ('CD', 'Estimated Resident Population (people) (reference year)'),
                    ('CE', 'Estimated Resident Population (people) (current year)'),
                    ('CF-CO', 'Growth in Estimated Resident Population (percent) (by year)'),
                    ('CP', 'Estimated Resident Population (State) (people) (reference year)'),
                    ('CQ', 'Estimated Resident Population (State) (people) (current year)'),
                    ('CR-DA', 'Growth in Estimated Resident (State) Population (percent) (by year)'),
                    
                    ('DB', 'Population of age 0-14 years (percent) (current year)'),
                    ('DC', 'Population of age 15-24 years (percent) (current year)'),
                    ('DD', 'Population of age 25-64 years (percent) (current year)'),
                    ('DE', 'Population of age 65-84 years (percent) (current year)'),
                    ('DF', 'Population of age 85+ years (percent) (current year)'),
                    ('DG', 'Population of age 0-14 years (State) (percent) (current year)'),
                    ('DH', 'Population of age 15-24 years (State) (percent) (current year)'),
                    ('DI', 'Population of age 25-64 years (State) (percent) (current year)'),
                    ('DJ', 'Population of age 65-84 years (State) (percent) (current year)'),
                    ('DK', 'Population of age 85+ years (State) (percent) (current year)'),
                    
                    ('DL', 'Proportion of people who speak language other than English at home (percent) (reference year)'),
                    ('DM', 'Proportion of people who speak language other than English at home (percent) (current year)'),
                    ('DN', 'Proportion of people who speak language other than English at home (rank) (reference year)'),
                    ('DO', 'Proportion of people who speak language other than English at home (rank) (current year)'),
                    ('DP', 'Proportion of people who speak language other than English at home (State) (percent) (reference year)'),
                    ('DQ', 'Proportion of people who speak language other than English at home (State) (percent) (current year)'),

                    ('DR', 'Persons on pensions/allowances (per mille) (reference year)'),
                    ('DS', 'Persons on pensions/allowances (per mille) (current year)'),
                    ('DT', 'Persons on pensions/allowances (rank) (reference year)'),
                    ('DU', 'Persons on pensions/allowances (rank) (current year)'),
                    ('DV', 'Persons on pensions/allowances (State) (per mille) (reference year)'),
                    ('DW', 'Persons on pensions/allowances (State) (per mille) (current year)'),
            ],
            "rows": [
                    ('1', 'Column headers (main)'),
                    ('2', 'Column headers (years)'),
                    ('3', 'Column headers (units)'),
                    ('4...', 'One row per SA4 region'),
            ],
            "notes": ["Contains raw SA4-level data for all widgets"],
        },
        {
            "name": "Society",
            "cols": [
                    ('A', 'Row identifier'),
                    ('C', 'Life Expectancy (years) (reference year)'),
                    ('D', 'Life Expectancy (years) (current year)'),
                    ('G', 'Adults who are overweight or obese (%) (reference year)'),
                    ('H', 'Adults who are overweight or obese (%) (current year)'),
                    ('K', 'Homeless persons per 10k population (persons) (reference year)'),
                    ('L', 'Homeless persons per 10k population (persons) (current year)'),
                    ('O', 'Proportion of households that own their own home (percentage) (reference year)'),
                    ('P', 'Proportion of households that own their own home (percentage) (current year)'),
                    ('S', 'People with vocational or higher education qualification (percentage) (reference year)'),
                    ('T', 'People with vocational or higher education qualification (percentage) (current year)'),
            ],
            "rows": [
                    ('Australia', 'Data for all-of-Australia'), 
                    ('...', 'All other rows ignored.'), 
            ],
            "notes": ["Some calculated/derived data and the all-Australia data for 'Society' metrics"],
        },
        {
            "name": "Environment",
            "cols": [
                    ('A', 'Row identifier'),
                    ('C', 'Terrestrial area protected (percent) (reference year)'),
                    ('D', 'Terrestrial area protected (percent) (current year)'),
                    ('G', 'Domestic involving nature activity trip (visits per resident) (reference year)'),
                    ('H', 'Domestic involving nature activity trip (visits per resident) (current year)'),
            ],
            "rows": [
                    ('Australia', 'Data for all-of-Australia'), 
                    ('...', 'All other rows ignored.'), 
            ],
            "notes": ["Some calculated/derived data and the all-Australia data for 'Environment' metrics"],
        },
        {
            "name": "Economy",
            "cols": [
                    ('A', 'Row identifier'),
                    ('C', 'Labour force participation (percent) (reference year)'),
                    ('D', 'Labour force participation (percent) (current year)'),
                    ('G', 'Proportion who are earning or learning (percent) (reference year)'),
                    ('H', 'Proportion who are earning or learning (percent) (current year)'),
                    ('K', 'Real median weekly household income ($) (reference year)'),
                    ('L', 'Real median weekly household income ($) (current year)'),
                    ('O', 'Ratio (P10:P50) (reference year)'),
                    ('P', 'Ratio (P10:P50) (current year)'),
                    ('S', 'New business entry rate (percent) (reference year)'),
                    ('T', 'New business entry rate (percent) (current year)'),
            ],
            "rows": [
                    ('Australia', 'Data for all-of-Australia'), 
                    ('...', 'All other rows ignored.'), 
            ],
            "notes": ["Some calculated/derived data and the all-Australia data for 'Economy' metrics"],
        },
        {
            "name": "Governance",
            "cols": [
                    ('A', 'Row identifier'),
                    ('C', 'Overseas-born Australian residents who are Australian citizens (percent) (reference year)'),
                    ('D', 'Overseas-born Australian residents who are Australian citizens (percent) (current year)'),
            ],
            "rows": [
                    ('Australia', 'Data for all-of-Australia'), 
                    ('...', 'All other rows ignored.'), 
            ],
            "notes": ["Some calculated/derived data and the all-Australia data for 'Governance' metrics"],
        },
        {
            "name": "ERP",
            "cols": [
                    ('A', 'Row identifier'),
                    ('C', 'Estimated Resident Population (people) (reference year)'),
                    ('D', 'Estimated Resident Population (people) (current year)'),
                    ('E-N', 'Growth in Estimated Resident Population (percent) (by year)'),
            ],
            "rows": [
                    ('Australia', 'Data for all-of-Australia'), 
                    ('...', 'All other rows ignored.'), 
            ],
            "notes": ["Some calculated/derived data and the all-Australia data for 'ERP' metrics"],
        },
        {
            "name": "Pop by age",
            "cols": [
                    ('A', 'Row identifier'),
                    ('C', 'Population of age 0-14 years (percent) (current year)'),
                    ('D', 'Population of age 15-24 years (percent) (current year)'),
                    ('E', 'Population of age 25-64 years (percent) (current year)'),
                    ('F', 'Population of age 65-84 years (percent) (current year)'),
                    ('G', 'Population of age 85+ years (percent) (current year)'),
            ],
            "rows": [
                    ('Australia', 'Data for all-of-Australia'), 
                    ('...', 'All other rows ignored.'), 
            ],
            "notes": ["Some calculated/derived data and the all-Australia data for 'Pop by Age' metrics"],
        },
        {
            "name": "context",
            "cols": [
                    ('A', 'Row identifier'),
                    ('C', 'Proportion of people who speak language other than English at home (percent) (reference year)'),
                    ('D', 'Proportion of people who speak language other than English at home (percent) (current year)'),
                    ('G', 'Persons on pensions/allowances (per mille) (reference year)'),
                    ('H', 'Persons on pensions/allowances (per mille) (current year)'),
            ],
            "rows": [
                    ('Australia', 'Data for all-of-Australia'), 
                    ('...', 'All other rows ignored.'), 
            ],
            "notes": ["Some calculated/derived data and the all-Australia data for 'Context' metrics"],
        },
        {
            "name": "Explanatory note",
            "cols": [],
            "rows": [],
            "notes": ["For human use only - not read by uploader"],
        },
    ], 
}

class PercentRanker(object):
    def __init__(self, lst):
        self._list = sorted(filter(lambda x : x != "#", lst))
        self._n = len(self._list)
    def pct_rank(self, val):
        for pos in range(0, self._n): 
            if val <= self._list[pos]:
                break
        return float(pos+1)/float(self._n+1)*100.0

def find_australia_row(sheet):
    for row in range(1,sheet.max_row+1):
        if sheet["A%d" % row].value == "Australia":
            return row
    raise LoaderException("No Australia row in sheet %s" % sheet.title)

def format_year(yin):
    if isinstance(yin, int):
        yout = unicode(yin)
    elif re.search(r"[^0-9]+", yin):
        # Financial Year for cleaning
        bits = re.split(r"[^0-9]+", yin, maxsplit=1)
        yout = "%d/%02d" % (int(bits[0]), int(bits[1]))
    else:
        yout = yin
    return yout

def std_init(wb, wurl, sa_rcol, sa_ccol, au_sheet, au_rcol, au_ccol):
    sa4_sheet = wb["SA4"]
    aus_sheet = wb[au_sheet]
    data = {}
    data["ref_year"] = format_year(sa4_sheet["%s2" % sa_rcol].value)
    data["curr_year"] = format_year(sa4_sheet["%s2" % sa_ccol].value)
    data["ranker_ry"] = PercentRanker([sa4_sheet["%s%d" % (sa_rcol, row)].value for row in range(4, sa4_sheet.max_row+1)])
    data["ranker_cy"] = PercentRanker([sa4_sheet["%s%d" % (sa_ccol, row)].value for row in range(4, sa4_sheet.max_row+1)])
    ozrow = find_australia_row(aus_sheet)
    data["aus_ry"] = aus_sheet["%s%d" % (au_rcol, ozrow)].value
    data["aus_cy"] = aus_sheet["%s%d" % (au_ccol, ozrow)].value
    data["gmain"] = get_graph(wurl, wurl + "_param", wurl + "_change")
    data["grank"] = get_graph(wurl, wurl + "_param", wurl + "_ranking")
    return data

def std_load(ws, row, state_row, pval, data, wurl, rcol, ccol, st_rcol, st_ccol):
    ry = ws["%s%d" % (rcol, row)].value
    cy = ws["%s%d" % (ccol, row)].value
    if ry == "#":
        return
    set_statistic_data(wurl, wurl + "_param", wurl + "_region", cy, 
                                    pval=pval,
                                    trend=cmp(cy, ry))
    set_statistic_data(wurl, wurl + "_param", wurl + "_region_old", ry, 
                                    pval=pval)
    set_statistic_data(wurl, wurl + "_param", wurl + "_aus", data["aus_cy"], 
                                    pval=pval)
    clear_graph_data(data["gmain"], pval=pval)
    add_graph_data(data["gmain"], "reference", pval=pval, cluster="region", value=ry)
    add_graph_data(data["gmain"], "reference", pval=pval, cluster="state", value=ws["%s%d" % (st_rcol, state_row)].value)
    add_graph_data(data["gmain"], "reference", pval=pval, cluster="australia", value=data["aus_ry"])
    add_graph_data(data["gmain"], "current", pval=pval, cluster="region", value=cy)
    add_graph_data(data["gmain"], "current", pval=pval, cluster="state", value=ws["%s%d" % (st_ccol, state_row)].value)
    add_graph_data(data["gmain"], "current", pval=pval, cluster="australia", value=data["aus_cy"])
    set_dataset_override(data["gmain"], "reference", data["ref_year"])
    set_dataset_override(data["gmain"], "current", data["curr_year"])
    clear_graph_data(data["grank"], pval=pval)
    add_graph_data(data["grank"], "ranking", pval=pval, horiz_value=data["ref_year"].split("/")[0], 
                        value=data["ranker_ry"].pct_rank(ry))
    add_graph_data(data["grank"], "ranking", pval=pval, horiz_value=data["curr_year"].split("/")[0], 
                        value=data["ranker_cy"].pct_rank(cy))
    set_actual_frequency_display_text(wurl, wurl + "_param", 
                        "%s-%s" % (data["ref_year"], data["curr_year"]), 
                        pval=pval)
    return

def init_life_expectancy(wb):
    return std_init(wb, "life_expectancy", "C", "D", "Society", "C", "D")

def load_life_expectancy(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "life_expectancy", "C", "D", "X", "Y")
    return

def init_obesity(wb):
    return std_init(wb, "life_expectancy", "G", "H", "Society", "G", "H")

def load_obesity(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "obesity", "G", "H", "Z", "AA")
    return

def init_homelessness(wb):
    return std_init(wb, "homelessness", "K", "L", "Society", "K", "L")

def load_homelessness(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "homelessness", "K", "L", "AB", "AC")
    return

def init_home_ownership(wb):
    return std_init(wb, "home_ownership", "O", "P", "Society", "O", "P")

def load_home_ownership(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "home_ownership", "O", "P", "AD", "AE")
    return

def init_higher_edu(wb):
    return std_init(wb, "higher_edu", "S", "T", "Society", "S", "T")

def load_higher_edu(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "higher_edu", "S", "T", "AF", "AG")
    return

def init_labour_force(wb):
    return std_init(wb, "labour_force", "AH", "AI", "Economy", "C", "D")

def load_labour_force(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "labour_force", "AH", "AI", "BB", "BC")
    return

def init_earn_learn(wb):
    return std_init(wb, "earn_learn", "AL", "AM", "Economy", "G", "H")

def load_earn_learn(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "earn_learn", "AL", "AM", "BD", "BE")
    return

def init_household_income(wb):
    return std_init(wb, "household_income", "AP", "AQ", "Economy", "K", "L")

def load_household_income(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "household_income", "AP", "AQ", "BF", "BG")
    return

def init_income_spread(wb):
    return std_init(wb, "income_spread", "AT", "AU", "Economy", "O", "P")

def load_income_spread(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "income_spread", "AT", "AU", "BH", "BI")
    return

def init_new_business(wb):
    return std_init(wb, "new_business", "AX", "AY", "Economy", "S", "T")

def load_new_business(ws, row, state_row, pval, data):
    std_load(ws, row, state_row, pval, data, "new_business", "AX", "AY", "BJ", "BK")
    return

def init_pop_age(wb):
    sa4_sheet = wb["SA4"]
    popage_sheet = wb["Pop by age"]
    data = {}
    data["year"] = sa4_sheet["DB1"].value
    data["graph"] = get_graph("population_age", "pop_age_param", "pop_age_graph")
    ozrow = find_australia_row(popage_sheet)
    data["aus_0_14"] = popage_sheet["C%d" % ozrow].value
    data["aus_15_24"] = popage_sheet["D%d" % ozrow].value
    data["aus_25_64"] = popage_sheet["E%d" % ozrow].value
    data["aus_65_84"] = popage_sheet["F%d" % ozrow].value
    data["aus_85plus"] = popage_sheet["G%d" % ozrow].value
    return data

def load_pop_age(ws, row, state_row, pval, data): 
    set_statistic_data("population_age", "pop_age_param", "yrs_0_14", ws["DB%d" % row].value, pval=pval)
    set_statistic_data("population_age", "pop_age_param", "yrs_15_24", ws["DC%d" % row].value, pval=pval)
    set_statistic_data("population_age", "pop_age_param", "yrs_25_64", ws["DD%d" % row].value, pval=pval)
    set_statistic_data("population_age", "pop_age_param", "yrs_65_84", ws["DE%d" % row].value, pval=pval)
    set_statistic_data("population_age", "pop_age_param", "yrs_85plus", ws["DF%d" % row].value, pval=pval)
    clear_graph_data(data["graph"], pval=pval)
    add_graph_data(data["graph"], pval=pval, cluster="region", dataset="yrs_0_14", 
                        value=ws["DB%d" % row].value)
    add_graph_data(data["graph"], pval=pval, cluster="region", dataset="yrs_15_24", 
                        value=ws["DC%d" % row].value)
    add_graph_data(data["graph"], pval=pval, cluster="region", dataset="yrs_25_64", 
                        value=ws["DD%d" % row].value)
    add_graph_data(data["graph"], pval=pval, cluster="region", dataset="yrs_65_84", 
                        value=ws["DE%d" % row].value)
    add_graph_data(data["graph"], pval=pval, cluster="region", dataset="yrs_85plus", 
                        value=ws["DF%d" % row].value)
    add_graph_data(data["graph"], pval=pval, cluster="state", dataset="yrs_0_14", 
                        value=ws["DG%d" % state_row].value)
    add_graph_data(data["graph"], pval=pval, cluster="state", dataset="yrs_15_24", 
                        value=ws["DH%d" % state_row].value)
    add_graph_data(data["graph"], pval=pval, cluster="state", dataset="yrs_25_64", 
                        value=ws["DI%d" % state_row].value)
    add_graph_data(data["graph"], pval=pval, cluster="state", dataset="yrs_65_84", 
                        value=ws["DJ%d" % state_row].value)
    add_graph_data(data["graph"], pval=pval, cluster="state", dataset="yrs_85plus", 
                        value=ws["DK%d" % state_row].value)
    add_graph_data(data["graph"], pval=pval, cluster="australia", dataset="yrs_0_14", 
                        value=data["aus_0_14"])
    add_graph_data(data["graph"], pval=pval, cluster="australia", dataset="yrs_15_24", 
                        value=data["aus_15_24"])
    add_graph_data(data["graph"], pval=pval, cluster="australia", dataset="yrs_25_64", 
                        value=data["aus_25_64"])
    add_graph_data(data["graph"], pval=pval, cluster="australia", dataset="yrs_65_84", 
                        value=data["aus_65_84"])
    add_graph_data(data["graph"], pval=pval, cluster="australia", dataset="yrs_85plus", 
                        value=data["aus_85plus"])
    set_actual_frequency_display_text("population_age", "pop_age_param", 
                        "%d" % data["year"], 
                        pval=pval)
    return

widget_loaders = [
        { "label": "life_expectancy", "init": init_life_expectancy, "load": load_life_expectancy},
        { "label": "obesity", "init": init_obesity, "load": load_obesity},
        { "label": "homelessness", "init": init_homelessness, "load": load_homelessness},
        { "label": "home_ownership", "init": init_home_ownership, "load": load_home_ownership},
        { "label": "higher_edu", "init": init_higher_edu, "load": load_higher_edu},
        { "label": "labour_force", "init": init_labour_force, "load": load_labour_force},
        { "label": "earn_learn", "init": init_earn_learn, "load": load_earn_learn},
        { "label": "household_income", "init": init_household_income, "load": load_household_income},
        { "label": "income_spread", "init": init_income_spread, "load": load_income_spread},
        { "label": "new_business", "init": init_new_business, "load": load_new_business},
        { "label": "pop_age", "init": init_pop_age, "load": load_pop_age},
]

def upload_file(uploader, fh, actual_freq_display=None, verbosity=0):
    messages = []
    try:
        wb = load_workbook(fh)

        data = {}
        for wl in widget_loaders:
            data[wl["label"]] = wl["init"](wb)

        sa4_sheet = wb["SA4"]
        for row in range(4, sa4_sheet.max_row+1):
            sa4_code = sa4_sheet["A%d" % row].value
            sa4_name = sa4_sheet["B%d" % row].value
            if sa4_sheet["X%d" % row].value:
                state_row = row
            try:
                pval = get_paramval("param_sa4", sa4_code=sa4_code)
            except LoaderException:
                if verbosity > 0:
                    messages.append("sa4code %d not parametised" % sa4_code)
                continue
            if verbosity > 2:
                messages.append("Loading SA4 %s" % sa4_name)

            for wl in widget_loaders:
                wl["load"](sa4_sheet, row, state_row, pval, data[wl["label"]])
    except LoaderException, e:
        raise e
    except Exception, e:
        raise LoaderException("Invalid file: %s" % unicode(e))
    return messages

